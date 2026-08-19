"""Deterministic export schema, replay identity, and workbook contracts.

Fixtures cover synthetic aligned curves, parser-owned raw provenance, and
nonempty stochastic replay state across strict JSON and workbook views.
Frozen R22 field presence and insertion order remain part of the JSON contract.
Run-log text is checked independently from workbook JSON cell encoding.
"""

from __future__ import annotations

import json
from dataclasses import replace
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from tests.support.model_cases import (
    dataset_project,
    final_fit_result,
    fit_candidate,
    prepared_data,
    project,
)

from xrr_fitter.io.export_log import run_log_bytes
from xrr_fitter.io.export_tables import (
    DatasetExportData,
    ExportReplayIdentity,
    batch_workbook_bytes,
    compatibility_workbook_bytes,
    dataset_json_bytes,
    dataset_workbook_bytes,
    json_text,
    parameters_csv_bytes,
)
from xrr_fitter.io.project_codec import project_to_dict
from xrr_fitter.io.xy import read_xy_bytes
from xrr_fitter.model.analysis import (
    McmcConfig,
    McmcReport,
    ParameterProfile,
    UncertaintyReport,
)
from xrr_fitter.model.data import PreparedData
from xrr_fitter.model.export import ExportFileRecord
from xrr_fitter.model.fitting import FitStageSummary
from xrr_fitter.model.instrument import PhysicsDiagnostic
from xrr_fitter.model.parameters import ParameterDefinition, ParameterValue
from xrr_fitter.model.project import OxideDecision

SOURCE_FIXTURE = Path(__file__).resolve().parents[2] / "fixtures/source/header_and_duplicates.xy"
PROJECT_REFERENCE = ExportFileRecord(
    "project_snapshot.xrrproj.json",
    123,
    "b" * 64,
)


def _context(dataset_id: str = "curve") -> DatasetExportData:
    data = prepared_data()
    model = data.intensity_normalized * 0.97
    residual = np.log10(model + data.r_floor) - np.log10(data.intensity_normalized + data.r_floor)
    source_candidate = fit_candidate()
    candidate = replace(
        source_candidate,
        parameters=(
            *source_candidate.parameters,
            ParameterValue("instrument.background", 2.5e-7, 0.0, 1.0e-5),
        ),
        qz_a_inv=data.qz_a_inv,
        model_normalized=model,
        log_residuals_decades=residual,
        weighted_residuals=residual / 0.05,
        sld_depth_a=np.array([0.0, 20.0, 40.0]),
        sld_profile_a2=np.array([0.0, 2.0e-5, 1.0e-5], dtype=complex),
    )
    result = final_fit_result(candidate)
    dataset = dataset_project(dataset_id, result=result)
    dataset = replace(
        dataset,
        source_path=data.source_path.as_posix(),
        source_sha256=data.source_sha256,
        fit_mask=tuple(bool(value) for value in data.fit_mask),
        fit_range_two_theta_deg=(
            float(data.two_theta_deg[0]),
            float(data.two_theta_deg[-1]),
        ),
    )
    value = project(dataset)
    return DatasetExportData(
        project=value,
        dataset=dataset,
        data=data,
        directory_mapping=((dataset_id, "001-curve-aaaaaaaa"),),
        selected=candidate,
        replay_identity=ExportReplayIdentity(1, 10101, 20202),
        matching_surface_oxide_rejection=False,
        project_reference=PROJECT_REFERENCE,
    )


def _context_with_parameter_metadata() -> DatasetExportData:
    original = _context()
    definitions = (
        ParameterDefinition(
            "scale",
            "Scale",
            "",
            "instrument",
            1.1,
            0.25,
            2.5,
            "linear",
            False,
            sharing_key="shared-scale",
        ),
        ParameterDefinition(
            "instrument.background",
            "Background",
            "reflectivity",
            "instrument",
            1.0e-7,
            0.0,
            1.0e-5,
            "linear",
            False,
            expert_only=True,
        ),
    )
    result = replace(
        original.result,
        parameter_definitions=definitions,
        warnings=("review-warning",),
    )
    dataset = replace(original.dataset, last_valid_result=result)
    value = replace(original.project, datasets=(dataset,))
    return DatasetExportData(
        project=value,
        dataset=dataset,
        data=original.data,
        directory_mapping=original.directory_mapping,
        selected=original.selected,
        replay_identity=original.replay_identity,
        matching_surface_oxide_rejection=original.matching_surface_oxide_rejection,
        project_reference=original.project_reference,
    )


def _context_with_length_parameter(dataset_id: str = "curve") -> DatasetExportData:
    original = _context(dataset_id)
    selected = replace(
        original.selected,
        parameters=(
            *original.selected.parameters,
            ParameterValue("component.0.thickness_a", 25.0, 2.0, 100.0),
        ),
    )
    result = replace(
        original.result,
        candidates=(selected,),
        warnings=("length-warning",),
    )
    dataset = replace(original.dataset, last_valid_result=result)
    value = replace(original.project, datasets=(dataset,))
    return DatasetExportData(
        project=value,
        dataset=dataset,
        data=original.data,
        directory_mapping=original.directory_mapping,
        selected=selected,
        replay_identity=original.replay_identity,
        matching_surface_oxide_rejection=original.matching_surface_oxide_rejection,
        project_reference=original.project_reference,
    )


def _project_contexts(*dataset_ids: str) -> tuple[DatasetExportData, ...]:
    originals = tuple(_context(dataset_id) for dataset_id in dataset_ids)
    value = project(*(context.dataset for context in originals))
    mapping = tuple(
        (dataset_id, f"{index:03d}-dataset-aaaaaaaa") for index, dataset_id in enumerate(dataset_ids, start=1)
    )
    return tuple(
        DatasetExportData(
            project=value,
            dataset=context.dataset,
            data=context.data,
            directory_mapping=mapping,
            selected=context.selected,
            replay_identity=context.replay_identity,
            matching_surface_oxide_rejection=(context.matching_surface_oxide_rejection),
            project_reference=context.project_reference,
        )
        for context in originals
    )


def _context_with_mcmc_diagnostics() -> DatasetExportData:
    original = _context()
    diagnostics = (
        PhysicsDiagnostic("review-code", "review diagnostic", (0, 2, 4)),
        PhysicsDiagnostic("surface_thin_layer_residual", "surface residual", (1,)),
    )
    uncertainty_diagnostic = PhysicsDiagnostic(
        "uncertainty-code",
        "uncertainty diagnostic",
        (3,),
    )
    selected = replace(original.selected, diagnostics=diagnostics)
    stage = FitStageSummary(
        stage="review-stage",
        candidate_ids=(selected.candidate_id,),
        best_objective=selected.objective,
        total_nfev=17,
        stop_reasons=("review-stop",),
    )
    mcmc = McmcReport(
        config=McmcConfig(walkers=2, burn_in=0, production_steps=2),
        child_seed=30303,
        parameter_names=("scale",),
        samples_physical=np.ones((2, 1)),
        log_probability=np.zeros(2),
        acceptance_fraction=np.full(2, 0.5),
        split_rhat=np.ones(1),
        effective_sample_size=np.ones(1),
        boundary_hits=(),
        candidate_id=selected.candidate_id,
    )
    uncertainty = UncertaintyReport(
        correlation_names=("scale",),
        correlation_matrix=np.ones((1, 1)),
        profiles=(
            ParameterProfile(
                "scale",
                np.array([0.9, 1.0, 1.1]),
                np.array([1.1, 1.0, 1.1]),
                True,
                True,
            ),
        ),
        bootstrap_intervals=(),
        bootstrap_failure_rate=0.0,
        boundary_hits=(),
        strong_correlations=(),
        systematic_residual=False,
        diagnostics=(uncertainty_diagnostic,),
        mcmc=mcmc,
        candidate_id=selected.candidate_id,
    )
    result = replace(
        original.result,
        candidates=(selected,),
        warnings=("review-warning",),
        child_seeds=(101, 202),
        stage_summaries=(stage,),
        uncertainty=uncertainty,
    )
    dataset = replace(
        original.dataset,
        last_valid_result=result,
        oxide_decisions=(OxideDecision("Si", "SiO2", "surface", False, "oxide-table-v1"),),
    )
    value = replace(original.project, datasets=(dataset,))
    return DatasetExportData(
        project=value,
        dataset=dataset,
        data=original.data,
        directory_mapping=original.directory_mapping,
        selected=selected,
        replay_identity=original.replay_identity,
        matching_surface_oxide_rejection=True,
        project_reference=original.project_reference,
    )


def _context_with_mismatched_uncertainty() -> DatasetExportData:
    original = _context_with_mcmc_diagnostics()
    selected = replace(
        original.selected,
        candidate_id="candidate-b",
        diagnostics=(),
    )
    result = replace(
        original.result,
        candidates=(original.selected, selected),
    )
    dataset = replace(original.dataset, last_valid_result=result)
    ui_state = replace(
        original.project.ui_state,
        selected_candidate_ids=((dataset.dataset_id, selected.candidate_id),),
    )
    value = replace(original.project, datasets=(dataset,), ui_state=ui_state)
    return replace(
        original,
        project=value,
        dataset=dataset,
        selected=selected,
    )


def _context_from_prepared_data(data: PreparedData) -> DatasetExportData:
    original = _context()
    selected = replace(
        original.selected,
        qz_a_inv=data.qz_a_inv,
        model_normalized=np.maximum(data.intensity_normalized, data.r_floor),
        log_residuals_decades=np.zeros(data.qz_a_inv.size),
        weighted_residuals=np.zeros(data.qz_a_inv.size),
    )
    result = replace(original.result, candidates=(selected,))
    dataset = replace(
        original.dataset,
        source_path=data.source_path.as_posix(),
        source_sha256=data.source_sha256,
        beam=data.beam,
        import_angle_offset_deg=data.import_angle_offset_deg,
        column_mapping=data.column_mapping,
        fit_mask=tuple(bool(value) for value in data.fit_mask),
        fit_range_two_theta_deg=(
            float(data.two_theta_deg[0]),
            float(data.two_theta_deg[-1]),
        ),
        last_valid_result=result,
    )
    value = replace(original.project, datasets=(dataset,))
    return DatasetExportData(
        project=value,
        dataset=dataset,
        data=data,
        directory_mapping=original.directory_mapping,
        selected=selected,
        replay_identity=original.replay_identity,
        matching_surface_oxide_rejection=False,
        project_reference=original.project_reference,
    )


def test_export_dataset_workbook_has_complete_aligned_sheets() -> None:
    context = _context()

    first = dataset_workbook_bytes(context)
    second = dataset_workbook_bytes(context)

    workbook = pd.ExcelFile(BytesIO(first))
    raw_data = pd.read_excel(BytesIO(first), sheet_name="RawData")
    model = pd.read_excel(BytesIO(first), sheet_name="ModelResiduals")
    required_raw_columns = {
        "row_index",
        "raw_row_index",
        "raw_row_text",
        "raw_parse_status",
        "source_row_group_json",
        "source_path",
        "source_sha256",
        "beam_json",
        "column_mapping_json",
        "normalization",
        "r_floor",
        "fit_ready",
        "warnings_json",
    }
    observed_layout = {
        "deterministic": first == second,
        "sheets": workbook.sheet_names,
        "raw_rows": len(raw_data),
        "model_rows": len(model),
        "raw_columns_present": required_raw_columns <= set(raw_data.columns),
        "fit_mask": model["fit_included"].tolist(),
    }
    expected_layout = {
        "deterministic": True,
        "sheets": [
            "Parameters",
            "Candidates",
            "RawData",
            "ModelResiduals",
            "Correlation",
            "Profiles",
            "RunInfo",
        ],
        "raw_rows": len(context.data.raw_rows),
        "model_rows": context.data.two_theta_deg.size,
        "raw_columns_present": True,
        "fit_mask": list(context.dataset.fit_mask),
    }
    assert observed_layout == expected_layout
    first_raw = raw_data.iloc[0]
    observed_provenance = {
        "source_path": first_raw["source_path"],
        "source_sha256": first_raw["source_sha256"],
        "beam": json.loads(first_raw["beam_json"]),
        "angle_offset": first_raw["import_angle_offset_deg"],
        "column_mapping": json.loads(first_raw["column_mapping_json"]),
        "normalization": first_raw["normalization"],
        "r_floor": first_raw["r_floor"],
        "fit_ready": bool(first_raw["fit_ready"]),
        "warnings": json.loads(first_raw["warnings_json"]),
    }
    expected_provenance = {
        "source_path": str(context.data.source_path),
        "source_sha256": context.data.source_sha256,
        "beam": payload_dataset_beam(context),
        "angle_offset": context.data.import_angle_offset_deg,
        "column_mapping": payload_column_mapping(context),
        "normalization": context.data.normalization,
        "r_floor": context.data.r_floor,
        "fit_ready": context.data.fit_ready,
        "warnings": list(context.data.warnings),
    }
    assert observed_provenance == expected_provenance


def test_export_workbook_preserves_real_parser_raw_provenance() -> None:
    source_bytes = SOURCE_FIXTURE.read_bytes()
    data = read_xy_bytes(
        source_bytes,
        source_path=SOURCE_FIXTURE.name,
        beam=_context().dataset.beam,
    )
    context = _context_from_prepared_data(data)

    raw_data = pd.read_excel(
        BytesIO(dataset_workbook_bytes(context)),
        sheet_name="RawData",
    )

    assert len(data.raw_rows) == 9
    assert data.two_theta_deg.size == 4
    assert len(raw_data) == len(data.raw_rows)
    assert raw_data["raw_row_text"].tolist() == list(data.raw_rows)
    assert raw_data["raw_parse_status"].tolist() == list(data.raw_parse_status)
    assert raw_data["row_index"].dropna().astype(int).tolist() == [0, 1, 2, 3]
    assert [json.loads(value) for value in raw_data["source_row_group_json"].dropna()] == [[3], [4, 5], [7], [8]]

    metadata_columns = (
        "source_path",
        "source_sha256",
        "beam_json",
        "import_angle_offset_deg",
        "column_mapping_json",
        "normalization",
        "r_floor",
        "fit_ready",
        "warnings_json",
    )
    assert {column: int(raw_data[column].notna().sum()) for column in metadata_columns} == {
        column: 1 for column in metadata_columns
    }
    first = raw_data.iloc[0]
    assert {
        "source_path": first["source_path"],
        "source_sha256": first["source_sha256"],
        "beam": json.loads(first["beam_json"]),
        "angle_offset": first["import_angle_offset_deg"],
        "column_mapping": json.loads(first["column_mapping_json"]),
        "normalization": first["normalization"],
        "r_floor": first["r_floor"],
        "fit_ready": bool(first["fit_ready"]),
        "warnings": json.loads(first["warnings_json"]),
    } == {
        "source_path": str(data.source_path),
        "source_sha256": data.source_sha256,
        "beam": payload_dataset_beam(context),
        "angle_offset": data.import_angle_offset_deg,
        "column_mapping": payload_column_mapping(context),
        "normalization": data.normalization,
        "r_floor": data.r_floor,
        "fit_ready": data.fit_ready,
        "warnings": list(data.warnings),
    }


def payload_dataset_beam(context: DatasetExportData) -> dict[str, object]:
    document = project_to_dict(context.project)
    return document["datasets"][0]["beam"]


def payload_column_mapping(context: DatasetExportData) -> dict[str, object]:
    document = project_to_dict(context.project)
    return document["datasets"][0]["column_mapping"]


def test_export_dataset_parameters_preserve_definition_metadata() -> None:
    context = _context_with_parameter_metadata()

    parameters = pd.read_excel(
        BytesIO(dataset_workbook_bytes(context)),
        sheet_name="Parameters",
    )

    scale = parameters.iloc[0]
    observed = {
        "columns": parameters.columns.tolist(),
        "name": scale["name"],
        "display_name": scale["display_name"],
        "category": scale["category"],
        "value": scale["value"],
        "lower": scale["lower"],
        "upper": scale["upper"],
        "transform": scale["transform"],
        "locked": bool(scale["locked"]),
        "integer": bool(scale["integer"]),
        "expert_only": bool(scale["expert_only"]),
        "sharing_key": scale["sharing_key"],
        "selected_candidate_id": scale["selected_candidate_id"],
        "background_expert_only": bool(parameters.iloc[1]["expert_only"]),
    }
    expected = {
        "columns": [
            "name",
            "display_name",
            "category",
            "value",
            "lower",
            "upper",
            "unit",
            "transform",
            "locked",
            "integer",
            "expert_only",
            "sharing_key",
            "selected_candidate_id",
        ],
        "name": "scale",
        "display_name": "Scale",
        "category": "instrument",
        "value": context.selected.parameters[0].value,
        "lower": context.selected.parameters[0].lower,
        "upper": context.selected.parameters[0].upper,
        "transform": "linear",
        "locked": False,
        "integer": False,
        "expert_only": False,
        "sharing_key": "shared-scale",
        "selected_candidate_id": context.selected.candidate_id,
        "background_expert_only": True,
    }
    assert observed == expected


def test_export_workbook_run_info_matches_json_and_keeps_strings_literal() -> None:
    context = _context("=1+1")
    payload = json.loads(dataset_json_bytes(context))
    workbook_bytes = dataset_workbook_bytes(context)
    row = pd.read_excel(BytesIO(workbook_bytes), sheet_name="RunInfo").iloc[0]

    assert row.index.tolist() == [
        "dataset_id",
        "dataset_directory",
        "dataset_directory_mapping",
        "source_path",
        "source_sha256",
        "schema_version",
        "algorithm_version",
        "project_master_seed",
        "service_seed_tree_version",
        "independent_root_child",
        "joint_root_child",
        "optimizer_child_seeds",
        "mcmc_child_seed",
        "selected_candidate_id",
        "uncertainty_absent_reason",
        "fitted_instrument_parameters",
        "confidence",
        "candidate_count",
        "warnings",
        "beam",
        "instrument",
        "scale_prior",
        "structure_evidence",
        "oxide_decisions",
        "fringe_screen_threshold_version",
        "budget_reclaim_threshold_version",
        "downsample_rule_version",
        "jacobian_version",
    ]
    assert row["dataset_id"] == "=1+1"
    assert json.loads(row["dataset_directory_mapping"]) == payload["run_info"]["dataset_directory_mapping"]
    assert json.loads(row["beam"]) == payload["run_info"]["beam"]
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    sheet = workbook["RunInfo"]
    columns = {cell.value: cell.column for cell in sheet[1]}
    dataset_cell = sheet.cell(row=2, column=columns["dataset_id"])
    assert dataset_cell.value == "=1+1"
    assert dataset_cell.data_type == "s"
    assert dataset_cell.hyperlink is None


def test_export_json_and_workbook_retain_nonempty_mcmc_replay_identity() -> None:
    context = _context_with_mcmc_diagnostics()
    payload = json.loads(dataset_json_bytes(context))
    row = pd.read_excel(
        BytesIO(dataset_workbook_bytes(context)),
        sheet_name="RunInfo",
    ).iloc[0]
    expected_json = {
        "fit_config": project_to_dict(context.project)["fit_config"],
        "service_seed_tree_version": 1,
        "independent_root_child": 10101,
        "joint_root_child": 20202,
        "optimizer_child_seeds": [101, 202],
        "mcmc_child_seed": 30303,
        "fitted_instrument_parameters": {"instrument.background": 2.5e-7},
    }
    expected_workbook = {key: value for key, value in expected_json.items() if key != "fit_config"}

    json_info = payload["run_info"]
    observed_json = {key: json_info[key] for key in expected_json}
    observed_workbook = {
        "service_seed_tree_version": row["service_seed_tree_version"],
        "independent_root_child": row["independent_root_child"],
        "joint_root_child": row["joint_root_child"],
        "optimizer_child_seeds": json.loads(row["optimizer_child_seeds"]),
        "mcmc_child_seed": row["mcmc_child_seed"],
        "fitted_instrument_parameters": json.loads(row["fitted_instrument_parameters"]),
    }
    assert observed_json == expected_json
    assert observed_workbook == expected_workbook


def _assert_mismatched_uncertainty_json(context: DatasetExportData) -> None:
    info = json.loads(dataset_json_bytes(context))["run_info"]
    assert info["mcmc_child_seed"] is None
    assert "uncertainty candidate mismatch" in info["uncertainty_absent_reason"]


def _assert_mismatched_uncertainty_workbook(context: DatasetExportData) -> None:
    workbook_bytes = dataset_workbook_bytes(context)
    workbook = pd.ExcelFile(BytesIO(workbook_bytes))
    run_info = pd.read_excel(BytesIO(workbook_bytes), sheet_name="RunInfo").iloc[0]
    assert pd.read_excel(BytesIO(workbook_bytes), sheet_name="Correlation").empty
    assert pd.read_excel(BytesIO(workbook_bytes), sheet_name="Profiles").empty
    assert "uncertainty candidate mismatch" in run_info["uncertainty_absent_reason"]
    assert workbook.sheet_names == [
        "Parameters",
        "Candidates",
        "RawData",
        "ModelResiduals",
        "Correlation",
        "Profiles",
        "RunInfo",
    ]


def _assert_mismatched_uncertainty_log(context: DatasetExportData) -> None:
    log = run_log_bytes(context).decode("utf-8")
    assert "mcmc_child_seed: None" in log
    assert "uncertainty-code" not in log
    assert "uncertainty candidate mismatch" in log


def test_export_omits_uncertainty_owned_by_another_selected_candidate() -> None:
    context = _context_with_mismatched_uncertainty()
    assert context.selected.candidate_id == "candidate-b"
    assert context.result.uncertainty.candidate_id != context.selected.candidate_id
    _assert_mismatched_uncertainty_json(context)
    _assert_mismatched_uncertainty_workbook(context)
    _assert_mismatched_uncertainty_log(context)


def test_export_workbook_json_codec_rejects_unknown_objects() -> None:
    with pytest.raises(TypeError):
        json_text({"unsupported": object()})


def test_export_csv_and_compatibility_workbook_are_deterministic() -> None:
    context = _context_with_length_parameter()

    csv_first = parameters_csv_bytes(context)
    csv_second = parameters_csv_bytes(context)
    workbook_first = compatibility_workbook_bytes((context,))
    workbook_second = compatibility_workbook_bytes((context,))

    workbook = pd.ExcelFile(BytesIO(workbook_first))
    summary = pd.read_excel(BytesIO(workbook_first), sheet_name="Summary")
    parameters_nm = pd.read_excel(BytesIO(workbook_first), sheet_name="Parameters_nm")
    curves = pd.read_excel(BytesIO(workbook_first), sheet_name="Curves")
    observed = {
        "csv_deterministic": csv_first == csv_second,
        "csv_has_crlf": b"\r\n" in csv_first,
        "csv_header": csv_first.splitlines()[0],
        "workbook_deterministic": workbook_first == workbook_second,
        "sheets": workbook.sheet_names,
        "summary_columns": summary.columns.tolist(),
        "warnings": json.loads(summary.loc[0, "warnings"]),
        "length_columns": parameters_nm.columns.tolist(),
        "curve_columns": curves.columns.tolist(),
        "value_angstrom": parameters_nm.loc[0, "value_angstrom"],
        "value_nm": parameters_nm.loc[0, "value_nm"],
    }
    expected = {
        "csv_deterministic": True,
        "csv_has_crlf": False,
        "csv_header": b"parameter_name,value,lower,upper",
        "workbook_deterministic": True,
        "sheets": ["Summary", "Parameters_nm", "Curves"],
        "summary_columns": [
            "dataset_id",
            "confidence",
            "objective",
            "selected_candidate_id",
            "warnings",
        ],
        "warnings": ["length-warning"],
        "length_columns": [
            "dataset_id",
            "parameter_name",
            "value_angstrom",
            "value_nm",
        ],
        "curve_columns": [
            "dataset_id",
            "two_theta_deg",
            "qz_a_inv",
            "intensity_raw",
            "intensity_normalized",
            "model_normalized",
            "fit_included",
        ],
        "value_angstrom": 25.0,
        "value_nm": 2.5,
    }
    assert observed == expected
    np.testing.assert_allclose(curves["two_theta_deg"], context.data.two_theta_deg)
    np.testing.assert_allclose(curves["intensity_raw"], context.data.intensity_raw)
    np.testing.assert_allclose(curves["model_normalized"], context.selected.model_normalized)


def test_export_batch_parameters_put_dataset_identity_first() -> None:
    first, second = _project_contexts("first", "second")

    workbook = batch_workbook_bytes((second, first))
    parameters = pd.read_excel(BytesIO(workbook), sheet_name="Parameters")

    assert parameters.columns.tolist() == [
        "dataset_id",
        "parameter_name",
        "value",
        "lower",
        "upper",
    ]
    assert parameters["dataset_id"].tolist() == [
        "first",
        "first",
        "second",
        "second",
    ]


def test_export_batch_rejects_contexts_from_different_projects() -> None:
    with pytest.raises(ValueError, match="same project"):
        batch_workbook_bytes((_context("first"), _context("second")))


def test_export_batch_rejects_non_export_contexts() -> None:
    with pytest.raises(TypeError, match="DatasetExportData"):
        batch_workbook_bytes((object(),))


def test_export_batch_requires_every_project_dataset_once() -> None:
    first, second = _project_contexts("first", "second")

    with pytest.raises(ValueError, match="exactly once"):
        batch_workbook_bytes((first,))
    with pytest.raises(ValueError, match="exactly once"):
        batch_workbook_bytes((first, first))


def test_export_log_records_warnings_seed_tree_stages_and_diagnostics() -> None:
    context = _context_with_mcmc_diagnostics()

    text = run_log_bytes(context).decode("utf-8")

    expected = "\n".join(
        (
            "dataset_id: curve",
            "confidence: 可信",
            "candidate_count: 1",
            "project_master_seed: 1201",
            "service_seed_tree_version: 1",
            "independent_root_child: 10101",
            "joint_root_child: 20202",
            "optimizer_child_seeds: [101,202]",
            "mcmc_child_seed: 30303",
            "warning: review-warning",
            'stage review-stage: candidate_ids=["candidate-0"]; '
            'best_objective=1.0; total_nfev=17; stop_reasons=["review-stop"]',
            "review-code: review diagnostic; full_data_indices=[0,2,4]; "
            "qz_a_inv_range=[0.00711815248633,0.0355906540161]",
            "surface_thin_layer_residual: surface residual; "
            "full_data_indices=[1]; "
            "qz_a_inv_range=[0.0142362995519,0.0142362995519]",
            "uncertainty-code: uncertainty diagnostic; full_data_indices=[3]; "
            "qz_a_inv_range=[0.0284725557375,0.0284725557375]",
            "疑似缺失自然氧化层（此前已拒绝建议）",
            "",
        )
    )
    assert text == expected


def test_export_log_ignores_unrelated_surface_oxide_rejection() -> None:
    original = _context()
    selected = replace(
        original.selected,
        diagnostics=(
            PhysicsDiagnostic(
                "surface_thin_layer_residual",
                "surface residual",
                (1,),
            ),
        ),
    )
    result = replace(original.result, candidates=(selected,))
    dataset = replace(
        original.dataset,
        last_valid_result=result,
        oxide_decisions=(OxideDecision("Cu", "Cu2O", "surface", False, "unrelated-version"),),
    )
    value = replace(original.project, datasets=(dataset,))
    context = DatasetExportData(
        project=value,
        dataset=dataset,
        data=original.data,
        directory_mapping=original.directory_mapping,
        selected=selected,
        replay_identity=original.replay_identity,
        matching_surface_oxide_rejection=False,
        project_reference=original.project_reference,
    )

    text = run_log_bytes(context).decode("utf-8")

    assert "surface_thin_layer_residual" in text
    assert "疑似缺失自然氧化层（此前已拒绝建议）" not in text


def test_export_log_retains_stale_diagnostic_indices_without_indexing_them() -> None:
    original = _context()
    stale_index = original.data.qz_a_inv.size + 4
    diagnostic = PhysicsDiagnostic(
        "stale-index",
        "retained diagnostic",
        (0, stale_index),
    )
    selected = replace(original.selected, diagnostics=(diagnostic,))
    result = replace(original.result, candidates=(selected,))
    dataset = replace(original.dataset, last_valid_result=result)
    value = replace(original.project, datasets=(dataset,))
    context = DatasetExportData(
        project=value,
        dataset=dataset,
        data=original.data,
        directory_mapping=original.directory_mapping,
        selected=selected,
        replay_identity=original.replay_identity,
        matching_surface_oxide_rejection=original.matching_surface_oxide_rejection,
        project_reference=original.project_reference,
    )

    text = run_log_bytes(context).decode("utf-8")

    qz = float(original.data.qz_a_inv[0])
    assert f"full_data_indices=[0,{stale_index}]" in text
    assert f"qz_a_inv_range=[{qz:.12g},{qz:.12g}]" in text
